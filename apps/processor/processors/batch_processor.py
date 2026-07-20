"""
Batch Processing Engine — CA Copilot
Processes thousands of documents concurrently across folders, financial years, ZIP archives.
Generates per-document and consolidated Excel workbooks.
"""
import asyncio
import time
import uuid
from pathlib import Path
from typing import List, Optional, Dict
from loguru import logger

from models.schemas import (
    ExtractionResult, BatchJob, BatchJobStatus, BatchExtractionRequest
)
from processors.document_pipeline import DocumentPipeline
from models.schemas import ExtractionRequest
from processors.duplicate_detector import detect_duplicates

# In-memory job store (in production, persist to SQLite or Redis)
_jobs: Dict[str, BatchJob] = {}
_job_status: Dict[str, dict] = {}


async def submit_batch_job(request: BatchExtractionRequest) -> str:
    """Submit a new batch processing job and return the job ID."""
    job_id = str(uuid.uuid4())
    job = BatchJob(
        job_id=job_id,
        files=request.file_paths,
        output_dir=request.output_dir,
        total_files=len(request.file_paths),
    )
    _jobs[job_id] = job

    # Start processing in background
    asyncio.create_task(_run_batch(job_id, request))
    logger.info(f"Batch job {job_id} queued: {len(request.file_paths)} files")
    return job_id


async def get_job_status(job_id: str) -> Optional[BatchJobStatus]:
    job = _jobs.get(job_id)
    if not job:
        return None
    progress = (job.processed_files / max(job.total_files, 1)) * 100
    return BatchJobStatus(
        job_id=job_id,
        status=job.status,
        total_files=job.total_files,
        processed_files=job.processed_files,
        failed_files=job.failed_files,
        progress_percent=round(progress, 1),
        current_file=_job_status.get(job_id, {}).get("current_file", ""),
        started_at=job.started_at,
        completed_at=job.completed_at,
    )


async def _run_batch(job_id: str, request: BatchExtractionRequest):
    """Background task that processes files one by one."""
    from datetime import datetime
    from exporters.excel_exporter import ExcelExporter

    job = _jobs[job_id]
    job.status = "processing"
    job.started_at = datetime.now().isoformat()
    output_path = Path(request.output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    results: List[ExtractionResult] = []

    for file_path in request.file_paths:
        _job_status[job_id] = {"current_file": Path(file_path).name}
        logger.info(f"[Batch {job_id}] Processing: {file_path}")

        try:
            ext_request = ExtractionRequest(
                file_path=file_path,
                ocr_enabled=True,
                preprocess_image=True,
            )
            pipeline = DocumentPipeline(ext_request)
            result = await pipeline.run()

            # Export to individual Excel file
            out_file = str(output_path / result.output_filename)
            exporter = ExcelExporter(result, out_file)
            await exporter.generate()

            results.append(result)
            job.processed_files += 1
            job.results.append(result)

        except Exception as e:
            logger.error(f"[Batch {job_id}] Failed: {file_path} — {e}")
            job.failed_files += 1
            job.errors[file_path] = str(e)

    # ── Cross-document duplicate detection ────────────────────────────────
    if len(results) > 1:
        logger.info(f"[Batch {job_id}] Running cross-document duplicate detection on {len(results)} results")
        duplicates = detect_duplicates(results)
        for result in results:
            result.duplicate_matches = [
                d for d in duplicates
                if d.source_file == result.original_filename or d.matched_file == result.original_filename
            ]

    # ── Generate consolidated workbook ────────────────────────────────────
    if request.generate_consolidated and results:
        try:
            consolidated_path = str(output_path / _get_consolidated_filename(request))
            await _generate_consolidated_workbook(results, consolidated_path)
            logger.info(f"[Batch {job_id}] Consolidated workbook: {consolidated_path}")
        except Exception as e:
            logger.error(f"[Batch {job_id}] Consolidated workbook failed: {e}")

    job.status = "completed" if job.failed_files == 0 else "partial"
    job.completed_at = datetime.now().isoformat()
    logger.info(
        f"[Batch {job_id}] COMPLETE — "
        f"{job.processed_files} processed, {job.failed_files} failed"
    )


async def _generate_consolidated_workbook(results: List[ExtractionResult], output_path: str):
    """
    Generate a single consolidated Excel workbook for all documents in a batch.
    Creates an index sheet + per-document summary rows.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Document Index"

    # Header
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0F294A", end_color="0F294A", fill_type="solid")
    regular_font = Font(name="Segoe UI", size=9, color="334155")
    thin_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    cols = [
        "S.No", "File Name", "Document Type", "Vendor Name", "Customer Name",
        "Invoice No", "Invoice Date", "Grand Total", "GSTIN", "Confidence",
        "Status", "Audit Findings", "Duplicates", "Processing Time (ms)"
    ]

    ws.row_dimensions[1].height = 22
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ok_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    for row_i, result in enumerate(results, 2):
        h = result.header
        has_critical = any(o.severity in ("critical", "high") for o in result.audit_observations)
        has_duplicates = bool(result.duplicate_matches)
        status = "⚠ Review" if result.needs_manual_review else ("✓ Complete" if not has_critical else "✗ Alert")

        row_data = [
            row_i - 1,
            result.original_filename,
            result.document_type.replace("_", " ").title(),
            h.vendor_name or (result.bank_header.account_holder if result.bank_header else ""),
            h.customer_name,
            h.invoice_number,
            h.invoice_date,
            h.grand_total or (result.bank_header.closing_balance if result.bank_header else ""),
            h.vendor_gstin,
            f"{result.overall_confidence_score:.0f}%",
            status,
            len(result.audit_observations),
            len(result.duplicate_matches),
            result.processing_duration_ms,
        ]

        row_fill = alert_fill if has_critical else (warn_fill if result.needs_manual_review else ok_fill)
        ws.row_dimensions[row_i].height = 18

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=ci, value=val)
            cell.font = regular_font
            cell.border = cell_border
            if ci == 11:  # Status column
                cell.fill = row_fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    # Audit Findings Sheet
    ws_audit = wb.create_sheet("Audit Findings")
    ws_audit.cell(row=1, column=1, value="Consolidated Audit Findings").font = Font(name="Segoe UI", bold=True, size=14, color="0B2240")
    audit_cols = ["S.No", "File", "Category", "Severity", "Description", "Evidence", "Legal Reference", "Recommendation", "Amount"]
    for ci, col_name in enumerate(audit_cols, 1):
        cell = ws_audit.cell(row=3, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    audit_row = 4
    sno = 1
    for result in results:
        for obs in result.audit_observations:
            row_data = [
                sno, result.original_filename, obs.category, obs.severity.upper(),
                obs.description, obs.evidence, obs.legal_reference, obs.recommendation, obs.amount_involved
            ]
            sev_fill = {
                "critical": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
                "high": alert_fill,
                "medium": warn_fill,
                "low": ok_fill,
                "info": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
            }.get(obs.severity, ok_fill)

            for ci, val in enumerate(row_data, 1):
                cell = ws_audit.cell(row=audit_row, column=ci, value=val)
                cell.font = regular_font
                cell.border = cell_border
                if ci == 4:
                    cell.fill = sev_fill

            audit_row += 1
            sno += 1

    # Duplicate Findings Sheet
    ws_dup = wb.create_sheet("Duplicate Alerts")
    ws_dup.cell(row=1, column=1, value="Duplicate Invoice Alerts").font = Font(name="Segoe UI", bold=True, size=14, color="0B2240")
    dup_cols = ["Source File", "Matched File", "Inv# Source", "Inv# Match", "Vendor Source", "Vendor Match", "Amt Source", "Amt Match", "Match Score", "Risk", "Reasons"]
    for ci, col_name in enumerate(dup_cols, 1):
        cell = ws_dup.cell(row=3, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    dup_row = 4
    seen_dup_pairs = set()
    for result in results:
        for dup in result.duplicate_matches:
            pair_key = tuple(sorted([dup.source_file, dup.matched_file]))
            if pair_key in seen_dup_pairs:
                continue
            seen_dup_pairs.add(pair_key)

            row_data = [
                dup.source_file, dup.matched_file,
                dup.source_invoice_number, dup.matched_invoice_number,
                dup.source_vendor, dup.matched_vendor,
                dup.source_amount, dup.matched_amount,
                f"{dup.match_score:.1f}%", dup.risk_level.upper(),
                "; ".join(dup.match_reasons),
            ]
            for ci, val in enumerate(row_data, 1):
                ws_dup.cell(row=dup_row, column=ci, value=val).font = regular_font

            dup_row += 1

    wb.save(output_path)
    logger.info(f"Consolidated workbook saved: {output_path}")


def _get_consolidated_filename(request: BatchExtractionRequest) -> str:
    from datetime import datetime
    client = (request.client_name or "Batch").replace(" ", "_")
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"{client}_Consolidated_{date_str}.xlsx"
