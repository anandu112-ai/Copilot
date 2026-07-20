"""
Audit Reports Exporter for CA Copilot Phase 3.
Generates styled Excel, CSV, and PDF files for Audit Findings, Fraud Indicators, GST Risks, Ledger Consistency, and Vendor Risk Profiles.
"""
import csv
import io
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
from loguru import logger

class AuditReportsExporter:
    def __init__(self, client_id: str, client_name: str):
        self.client_id = client_id
        self.client_name = client_name
        self._init_styles()

    def _init_styles(self):
        ff = "Segoe UI"
        self.f_title = Font(name=ff, size=16, bold=True, color="0F294A")
        self.f_subtitle = Font(name=ff, size=10, italic=True, color="475569")
        self.f_header = Font(name=ff, size=10, bold=True, color="FFFFFF")
        self.f_section = Font(name=ff, size=11, bold=True, color="0F294A")
        self.f_bold = Font(name=ff, size=9, bold=True, color="1E293B")
        self.f_reg = Font(name=ff, size=9, color="334155")
        
        self.fill_header = PatternFill(start_color="0F294A", end_color="0F294A", fill_type="solid")
        self.fill_section = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        self.fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        self.fill_crit = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Red
        self.fill_high = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")   # Orange
        self.fill_med = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")    # Amber
        self.fill_low = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")    # Slate

        thin = Side(border_style="thin", color="CBD5E1")
        double = Side(border_style="double", color="1E293B")
        self.border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)

    def export_excel(self, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Export audit findings report to styled Excel workbook."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AI Audit Observations"
        ws.views.sheetView[0].showGridLines = True

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"AI Audit Intelligence Findings — {self.client_name}"
        ws["A1"].font = self.f_title
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | CA Copilot isolated sandbox"
        ws["A2"].font = self.f_subtitle
        ws["A2"].alignment = Alignment(horizontal="center")

        # Table Header
        headers = ["Severity", "Title", "Category", "Description", "Evidence", "Statutory Rule Ref", "Impact Value"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = self.f_header
            cell.fill = self.fill_header

        row_num = 6
        for f in findings:
            sev_cell = ws.cell(row=row_num, column=1, value=f.get("severity", "").upper())
            sev_cell.font = self.f_bold
            
            sev = f.get("severity", "").lower()
            if sev == "critical":
                sev_cell.fill = self.fill_crit
            elif sev == "high":
                sev_cell.fill = self.fill_high
            elif sev == "medium":
                sev_cell.fill = self.fill_med
            else:
                sev_cell.fill = self.fill_low

            ws.cell(row=row_num, column=2, value=f.get("title")).font = self.f_bold
            ws.cell(row=row_num, column=3, value=f.get("category")).font = self.f_reg
            ws.cell(row=row_num, column=4, value=f.get("description")).font = self.f_reg
            ws.cell(row=row_num, column=5, value=f.get("evidence")).font = self.f_reg
            ws.cell(row=row_num, column=6, value=f.get("legal_reference")).font = self.f_reg
            ws.cell(row=row_num, column=7, value=f.get("impact_amount")).font = self.f_bold

            for col_idx in range(1, 8):
                ws.cell(row=row_num, column=col_idx).border = self.border_cell
            row_num += 1

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

        wb.save(output_path)
        logger.info(f"Audit Excel exported: {output_path}")
        return output_path

    def export_csv(self, findings: List[Dict[str, Any]]) -> str:
        """Export findings to a CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["AI Audit Intelligence Findings Report", f"Client: {self.client_name}"])
        writer.writerow([])
        writer.writerow(["Severity", "Title", "Category", "Description", "Evidence", "Statutory Rule Ref", "Impact Value", "Status"])
        for f in findings:
            writer.writerow([
                f.get("severity"),
                f.get("title"),
                f.get("category"),
                f.get("description"),
                f.get("evidence"),
                f.get("legal_reference"),
                f.get("impact_amount"),
                f.get("status")
            ])
        return output.getvalue()

    def export_pdf(self, findings: List[Dict[str, Any]], output_path: str) -> str:
        """Export findings to a styled PDF using PyMuPDF."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        doc = fitz.open()

        page = doc.new_page()
        width, height = page.rect.width, page.rect.height

        # Title bar
        page.draw_rect(fitz.Rect(0, 0, width, 80), fill=(15/255, 41/255, 74/255), color=None)
        page.insert_text(fitz.Point(30, 45), "CA Copilot AI Audit Observations", fontsize=16, color=(1, 1, 1))
        page.insert_text(fitz.Point(30, 65), f"Client: {self.client_name} | Generated: {datetime.now().strftime('%Y-%m-%d')} | Private Audit Sandbox", fontsize=8, color=(0.8, 0.8, 0.8))

        y = 120
        page.insert_text(fitz.Point(30, y), f"AI Observations List ({len(findings)} Flagged Issues):", fontsize=11, color=(15/255, 41/255, 74/255))
        y += 25

        for idx, f in enumerate(findings, 1):
            if y > height - 120:
                page = doc.new_page()
                y = 50

            sev = f.get("severity", "").upper()
            title = f.get("title", "")
            
            # Severity Indicator Box
            box_color = (0.9, 0.1, 0.1) if sev == "CRITICAL" else (0.9, 0.5, 0.1) if sev == "HIGH" else (0.8, 0.8, 0.1)
            page.draw_rect(fitz.Rect(30, y-10, 80, y+5), fill=box_color, color=None)
            page.insert_text(fitz.Point(35, y), sev, fontsize=8, color=(1, 1, 1))
            
            # Observation Title
            page.insert_text(fitz.Point(120, y), f"{idx}. {title}", fontsize=10, color=(15/255, 41/255, 74/255))
            y += 18

            # Details
            page.insert_text(fitz.Point(40, y), f"Category: {f.get('category')} | Legal Reference: {f.get('legal_reference') or 'N/A'}", fontsize=8, color=(0.4, 0.4, 0.4))
            y += 12
            
            desc_lines = self._wrap_text(f.get("description", ""), 85)
            for line in desc_lines:
                page.insert_text(fitz.Point(40, y), line, fontsize=8, color=(0.2, 0.2, 0.2))
                y += 12

            if f.get("impact_amount", 0) > 0:
                page.insert_text(fitz.Point(40, y), f"Tax / Value Impact: INR {f.get('impact_amount'):,.2f}", fontsize=8, color=(0.6, 0.1, 0.1))
                y += 12
            
            page.draw_line(fitz.Point(30, y), fitz.Point(width - 30, y), color=(0.8, 0.8, 0.8))
            y += 20

        doc.save(output_path)
        doc.close()
        logger.info(f"Audit PDF exported: {output_path}")
        return output_path

    def _wrap_text(self, text: str, max_chars: int) -> List[str]:
        words = text.split(" ")
        lines = []
        curr = ""
        for w in words:
            if len(curr) + len(w) + 1 > max_chars:
                lines.append(curr)
                curr = w
            else:
                curr = f"{curr} {w}" if curr else w
        if curr:
            lines.append(curr)
        return lines
