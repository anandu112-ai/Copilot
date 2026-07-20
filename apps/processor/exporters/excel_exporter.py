"""
Excel Exporter — CA Copilot
Generates professional, CA-ready Excel workbooks from ExtractionResult objects.
Supports: Invoice/GST workbooks and Bank Statement workbooks.

Workbook structure (Invoice):
  Sheet 1: Invoice Summary     — header fields, vendor/buyer/bank details
  Sheet 2: Line Items          — product table with formulas and auto-filters
  Sheet 3: Tax Summary         — tax breakdown
  Sheet 4: Validation Results  — arithmetic, GSTIN, tax type checks
  Sheet 5: Extraction Confidence — per-field AI scores
  Sheet 6: Audit Observations  — compliance findings with legal refs
  Sheet 7: Document Metadata   — processing metadata

Workbook structure (Bank Statement):
  Sheet 1: Account Summary     — account header
  Sheet 2: Transactions        — full transaction ledger with audit flags
  Sheet 3: Category Summary    — pivot-style credit/debit by category
  Sheet 4: Validation Results  — balance reconciliation, 40A(3) checks
  Sheet 5: Audit Observations  — weekend txns, large txns, cash alerts
  Sheet 6: Document Metadata   — processing metadata
"""
from typing import List, Optional
from pathlib import Path
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from models.schemas import ExtractionResult


class ExcelExporter:
    def __init__(self, result: ExtractionResult, output_path: str):
        self.result = result
        self.output_path = output_path
        self._init_styles()

    def _init_styles(self):
        ff = "Segoe UI"
        # Fonts
        self.f_title   = Font(name=ff, size=15, bold=True, color="0B2240")
        self.f_header  = Font(name=ff, size=10, bold=True, color="FFFFFF")
        self.f_section = Font(name=ff, size=11, bold=True, color="0B2240")
        self.f_bold    = Font(name=ff, size=9, bold=True, color="1E293B")
        self.f_reg     = Font(name=ff, size=9, color="334155")
        self.f_small   = Font(name=ff, size=8, color="64748B")

        # Fills
        self.fill_header  = PatternFill(start_color="0F294A", end_color="0F294A", fill_type="solid")  # Navy
        self.fill_sub     = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # Dark Navy
        self.fill_section = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Slate-50
        self.fill_alt     = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Even lighter
        self.fill_pass    = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Green
        self.fill_warn    = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber
        self.fill_alert   = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Red
        self.fill_info    = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Blue
        self.fill_crit    = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")  # Deep red

        # Borders
        thin = Side(border_style="thin", color="CBD5E1")
        thick = Side(border_style="medium", color="94A3B8")
        double = Side(border_style="double", color="1E293B")
        self.border_cell  = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.border_total = Border(top=thick, bottom=double)
        self.border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    async def generate(self) -> str:
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()

        if self.result.document_type == "bank_statement":
            self._build_bank_workbook(wb)
        else:
            self._build_invoice_workbook(wb)

        wb.save(self.output_path)
        logger.info(f"Excel saved: {self.output_path}")
        return self.output_path

    # ════════════════════════════════════════════════════════════
    # INVOICE WORKBOOK
    # ════════════════════════════════════════════════════════════

    def _build_invoice_workbook(self, wb):
        ws1 = wb.active
        ws1.title = "Invoice Summary"
        self._sheet_invoice_summary(ws1)

        ws2 = wb.create_sheet("Line Items")
        self._sheet_line_items(ws2)

        ws3 = wb.create_sheet("Tax Summary")
        self._sheet_tax_summary(ws3)

        ws4 = wb.create_sheet("Validation Results")
        self._sheet_validation(ws4)

        ws5 = wb.create_sheet("Extraction Confidence")
        self._sheet_confidence(ws5)

        ws6 = wb.create_sheet("Audit Observations")
        self._sheet_audit_observations(ws6)

        ws7 = wb.create_sheet("Document Metadata")
        self._sheet_metadata(ws7)

    def _sheet_invoice_summary(self, ws):
        ws.views.sheetView[0].showGridLines = True
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 42

        h = self.result.header
        self._write_title(ws, 1, "CA Copilot — Invoice Summary")

        sections = [
            ("📄 Document Details", [
                ("Invoice Number",     h.invoice_number),
                ("Invoice Date",       h.invoice_date),
                ("Due Date",           h.due_date),
                ("Reference Number",   h.reference_number),
                ("Purchase Order No.", h.purchase_order_number),
                ("Document Type",      self.result.document_type.replace("_", " ").title()),
                ("Currency",           h.currency),
                ("Payment Terms",      h.payment_terms),
                ("Narration",          h.narration),
            ]),
            ("🏭 Vendor / Seller Details", [
                ("Vendor Name",        h.vendor_name),
                ("Vendor GSTIN",       h.vendor_gstin),
                ("Vendor PAN",         h.vendor_pan),
                ("Vendor Address",     h.vendor_address),
                ("Vendor State",       h.vendor_state),
                ("Vendor PIN",         h.vendor_pin),
            ]),
            ("👤 Buyer / Customer Details", [
                ("Customer Name",      h.customer_name),
                ("Customer GSTIN",     h.customer_gstin),
                ("Customer Address",   h.customer_address),
                ("Customer State",     h.customer_state),
                ("Customer PIN",       h.customer_pin),
                ("Place of Supply",    h.place_of_supply),
            ]),
            ("🚚 Transport & Logistics", [
                ("E-Way Bill Number",  h.eway_bill_number),
                ("Vehicle Number",     h.vehicle_number),
                ("Transporter Name",   h.transport_name),
                ("Transport Mode",     h.transport_mode),
                ("LR / GR Number",     h.lr_number),
            ]),
            ("🏦 Banking Details", [
                ("Bank Name",          h.bank_name),
                ("Account Number",     h.account_number),
                ("IFSC Code",          h.ifsc),
                ("UPI ID",             h.upi),
                ("SWIFT Code",         h.swift_code),
            ]),
            ("💰 Financial Summary", [
                ("Subtotal",           h.subtotal),
                ("Discount",           h.discount),
                ("Freight / Shipping", h.freight),
                ("Other Charges",      h.other_charges),
                ("Taxable Amount",     h.taxable_amount),
                ("CGST",               h.cgst),
                ("SGST",               h.sgst),
                ("IGST",               h.igst),
                ("CESS",               h.cess),
                ("TDS / TCS",          h.tds_tcs),
                ("Round Off",          h.round_off),
                ("Grand Total",        h.grand_total),
                ("Amount in Words",    h.amount_in_words),
            ]),
        ]

        row = 3
        for section_title, fields in sections:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value=section_title)
            c.font = self.f_section
            c.fill = self.fill_section
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1

            for label, val in fields:
                ws.row_dimensions[row].height = 18
                c1 = ws.cell(row=row, column=1, value=label)
                c2 = ws.cell(row=row, column=2, value=val or "—")
                c1.font = self.f_bold
                c2.font = self.f_reg
                c1.border = self.border_cell
                c2.border = self.border_cell
                c1.fill = self.fill_alt
                row += 1
            row += 1

    def _sheet_line_items(self, ws):
        headers = [
            "S.No", "Description", "HSN / SAC", "Quantity", "Unit",
            "Rate (₹)", "Discount", "Taxable Value (₹)",
            "CGST %", "CGST Amt (₹)", "SGST %", "SGST Amt (₹)",
            "IGST %", "IGST Amt (₹)", "Total (₹)", "Needs Review", "Review Reason"
        ]
        self._write_table_header(ws, 1, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        amt_cols = {6, 8, 10, 12, 14, 15}
        items = self.result.line_items

        for ri, item in enumerate(items, 2):
            ws.row_dimensions[ri].height = 18
            row_data = [
                item.sr_no, item.description, item.hsn_sac,
                self._to_float(item.quantity), item.unit,
                self._to_float(item.rate), self._to_float(item.discount),
                self._to_float(item.taxable_value),
                item.cgst_rate, self._to_float(item.cgst_amount),
                item.sgst_rate, self._to_float(item.sgst_amount),
                item.igst_rate, self._to_float(item.igst_amount),
                self._to_float(item.total),
                "YES" if item.needs_review else "NO",
                item.review_reason or "",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
                if ci in amt_cols and isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                if ci == 16:  # Needs Review
                    c.fill = self.fill_warn if val == "YES" else self.fill_pass
                    c.font = self.f_bold
                    c.alignment = Alignment(horizontal="center")
                if item.needs_review:
                    ws.row_dimensions[ri].height = 20

        # Total row
        if items:
            total_row = len(items) + 2
            ws.cell(row=total_row, column=2, value="TOTAL").font = self.f_bold
            sum_cols = [(8, "H"), (10, "J"), (12, "L"), (14, "N"), (15, "O")]
            for ci, col_l in sum_cols:
                c = ws.cell(row=total_row, column=ci,
                             value=f"=SUM({col_l}2:{col_l}{total_row - 1})")
                c.font = self.f_bold
                c.number_format = "#,##0.00"
                c.border = self.border_total
                c.alignment = Alignment(horizontal="right")

        self._autofit(ws)

    def _sheet_tax_summary(self, ws):
        self._write_title(ws, 1, "Tax Summary")
        h = self.result.header
        rows = [
            ("Taxable Value",       h.taxable_amount, False),
            ("CGST",                h.cgst,           False),
            ("SGST",                h.sgst,           False),
            ("IGST",                h.igst,           False),
            ("CESS",                h.cess,           False),
            ("TDS / TCS",           h.tds_tcs,        False),
            ("Round Off",           h.round_off,      False),
            ("Grand Total",         h.grand_total,    True),
            ("Amount in Words",     h.amount_in_words, False),
        ]
        for ri, (label, val, is_total) in enumerate(rows, 3):
            ws.row_dimensions[ri].height = 20
            c1 = ws.cell(row=ri, column=1, value=label)
            c2 = ws.cell(row=ri, column=2, value=self._to_float(val) if val else val or "—")
            c1.font = self.f_bold if is_total else self.f_reg
            c2.font = self.f_bold if is_total else self.f_reg
            if is_total:
                c1.border = self.border_total
                c2.border = self.border_total
            else:
                c1.border = self.border_cell
                c2.border = self.border_cell
            if isinstance(c2.value, (int, float)):
                c2.number_format = "#,##0.00"
                c2.alignment = Alignment(horizontal="right")
        self._autofit(ws)

    def _sheet_validation(self, ws):
        self._write_title(ws, 1, "Validation Results")
        headers = ["Checkpoint", "Logic / Rule", "Calculated Value", "Status", "Observation"]
        self._write_table_header(ws, 3, headers)
        ws.freeze_panes = "A4"

        status_fills = {
            "PASS": self.fill_pass,
            "WARNING": self.fill_warn,
            "HIGH ALERT": self.fill_alert,
            "INFO": self.fill_info,
        }

        for ri, chk in enumerate(self.result.validation_results, 4):
            ws.row_dimensions[ri].height = 22
            row_data = [chk.name, chk.logic, chk.result_value, chk.status, chk.observation]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
                c.alignment = Alignment(wrap_text=True, vertical="center")
            # Highlight status cell
            status_cell = ws.cell(row=ri, column=4)
            status_cell.fill = status_fills.get(chk.status, self.fill_info)
            status_cell.font = self.f_bold
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
        self._autofit(ws)

    def _sheet_confidence(self, ws):
        self._write_title(ws, 1, "Extraction Confidence")
        headers = ["Field Name", "Extracted Value", "Extraction Method", "Confidence %", "Needs Review"]
        self._write_table_header(ws, 3, headers)
        ws.freeze_panes = "A4"

        for ri, fc in enumerate(self.result.field_confidences, 4):
            ws.row_dimensions[ri].height = 20
            row_data = [
                fc.field_name.replace("_", " ").title(),
                fc.extracted_value or "—",
                fc.extraction_method.replace("_", " ").title(),
                fc.confidence_score,
                "YES" if fc.needs_review else "NO",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
            # Color confidence score
            conf_cell = ws.cell(row=ri, column=4)
            conf_cell.number_format = "0.0"
            score = fc.confidence_score
            if score >= 90:
                conf_cell.fill = self.fill_pass
            elif score >= 70:
                conf_cell.fill = self.fill_warn
            else:
                conf_cell.fill = self.fill_alert
            conf_cell.font = self.f_bold
            conf_cell.alignment = Alignment(horizontal="right")
            # Review column
            rev_cell = ws.cell(row=ri, column=5)
            rev_cell.fill = self.fill_warn if fc.needs_review else self.fill_pass
            rev_cell.font = self.f_bold
            rev_cell.alignment = Alignment(horizontal="center")
        self._autofit(ws)

    def _sheet_audit_observations(self, ws):
        self._write_title(ws, 1, "Audit Observations")
        headers = [
            "Category", "Severity", "Description",
            "Evidence", "Legal Reference", "Recommendation", "Amount Involved"
        ]
        self._write_table_header(ws, 3, headers)
        ws.freeze_panes = "A4"

        sev_fills = {
            "critical": self.fill_crit,
            "high": self.fill_alert,
            "medium": self.fill_warn,
            "low": self.fill_pass,
            "info": self.fill_info,
        }

        for ri, obs in enumerate(self.result.audit_observations, 4):
            ws.row_dimensions[ri].height = 35
            row_data = [
                obs.category, obs.severity.upper(), obs.description,
                obs.evidence, obs.legal_reference, obs.recommendation, obs.amount_involved
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
                c.alignment = Alignment(wrap_text=True, vertical="top")
            sev_cell = ws.cell(row=ri, column=2)
            sev_cell.fill = sev_fills.get(obs.severity, self.fill_info)
            sev_cell.font = self.f_bold
            sev_cell.alignment = Alignment(horizontal="center", vertical="center")

        if not self.result.audit_observations:
            ws.cell(row=4, column=1, value="✅ No audit observations found — document appears compliant").font = self.f_bold

        self._autofit(ws)

    def _sheet_metadata(self, ws):
        self._write_title(ws, 1, "Document Metadata")
        result = self.result
        h = result.header
        rows = [
            ("Original Filename",      result.original_filename),
            ("Output Filename",        result.output_filename),
            ("Document Type",          result.document_type.replace("_", " ").title()),
            ("Page Count",             str(result.page_count)),
            ("File Size",              f"{result.file_size_bytes:,} bytes"),
            ("OCR Used",               "Yes" if result.is_ocr_used else "No"),
            ("Image Preprocessed",     "Yes" if result.image_preprocessed else "No"),
            ("OCR Language",           result.ocr_language),
            ("Processing Duration",    f"{result.processing_duration_ms} ms"),
            ("Overall AI Confidence",  f"{result.overall_confidence_score:.1f}%"),
            ("Needs Manual Review",    "Yes" if result.needs_manual_review else "No"),
            ("Audit Observations",     str(len(result.audit_observations))),
            ("Validation Checks",      str(len(result.validation_results))),
            ("Processing Engine",      "CA Copilot v2.0 — AGY NLP"),
            ("Generated On",           "2026-07-20"),
        ]
        for ri, (label, val) in enumerate(rows, 3):
            ws.row_dimensions[ri].height = 18
            c1 = ws.cell(row=ri, column=1, value=label)
            c2 = ws.cell(row=ri, column=2, value=val)
            c1.font = self.f_bold
            c2.font = self.f_reg
            c1.border = self.border_cell
            c2.border = self.border_cell
            c1.fill = self.fill_alt
        self._autofit(ws)

    # ════════════════════════════════════════════════════════════
    # BANK STATEMENT WORKBOOK
    # ════════════════════════════════════════════════════════════

    def _build_bank_workbook(self, wb):
        ws1 = wb.active
        ws1.title = "Account Summary"
        self._sheet_bank_summary(ws1)

        ws2 = wb.create_sheet("Transactions")
        self._sheet_transactions(ws2)

        ws3 = wb.create_sheet("Category Summary")
        self._sheet_category_summary(ws3)

        ws4 = wb.create_sheet("Validation Results")
        self._sheet_validation(ws4)

        ws5 = wb.create_sheet("Audit Observations")
        self._sheet_audit_observations(ws5)

        ws6 = wb.create_sheet("Document Metadata")
        self._sheet_metadata(ws6)

    def _sheet_bank_summary(self, ws):
        bh = self.result.bank_header
        if not bh:
            ws.cell(row=1, column=1, value="No bank statement header extracted")
            return
        self._write_title(ws, 1, "Bank Account Summary")
        rows = [
            ("Account Holder",   bh.account_holder),
            ("Account Number",   bh.account_number),
            ("Bank Name",        bh.bank_name),
            ("Branch",           bh.branch),
            ("IFSC Code",        bh.ifsc),
            ("Account Type",     bh.account_type),
            ("Statement From",   bh.statement_from),
            ("Statement To",     bh.statement_to),
            ("Opening Balance",  bh.opening_balance),
            ("Closing Balance",  bh.closing_balance),
            ("Total Credits",    bh.total_credits),
            ("Total Debits",     bh.total_debits),
            ("Credit Count",     str(bh.total_credit_count)),
            ("Debit Count",      str(bh.total_debit_count)),
            ("Currency",         bh.currency),
        ]
        for ri, (label, val) in enumerate(rows, 3):
            ws.row_dimensions[ri].height = 18
            c1 = ws.cell(row=ri, column=1, value=label)
            c2 = ws.cell(row=ri, column=2, value=val or "—")
            c1.font = self.f_bold
            c2.font = self.f_reg
            c1.border = self.border_cell
            c2.border = self.border_cell
            c1.fill = self.fill_alt
        self._autofit(ws)

    def _sheet_transactions(self, ws):
        headers = [
            "S.No", "Date", "Value Date", "Description / Narration",
            "Reference No.", "Chq / UTR", "Debit (₹)", "Credit (₹)",
            "Balance (₹)", "Type", "Mode", "Category", "Counterparty",
            "Audit Flag", "Needs Review"
        ]
        self._write_table_header(ws, 1, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        amt_cols = {7, 8, 9}
        for ri, txn in enumerate(self.result.transactions, 2):
            ws.row_dimensions[ri].height = 18
            row_data = [
                ri - 1, txn.date, txn.value_date, txn.description,
                txn.reference_number, txn.chq_ref,
                self._to_float(txn.debit), self._to_float(txn.credit),
                self._to_float(txn.balance),
                (txn.transaction_type or "").upper(),
                txn.payment_mode or "",
                txn.category or "",
                txn.counterparty or "",
                txn.audit_flag or "",
                "YES" if txn.needs_review else "NO",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
                if ci in amt_cols and isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")

            # Color debit/credit cells
            debit_c = ws.cell(row=ri, column=7)
            credit_c = ws.cell(row=ri, column=8)
            if txn.transaction_type == "debit" and txn.debit:
                debit_c.fill = self.fill_alert
            elif txn.transaction_type == "credit" and txn.credit:
                credit_c.fill = self.fill_pass

            # Audit flag highlights
            if txn.audit_flag:
                ws.cell(row=ri, column=14).fill = self.fill_warn
            if txn.needs_review:
                ws.cell(row=ri, column=15).fill = self.fill_warn
                ws.cell(row=ri, column=15).font = self.f_bold

        self._autofit(ws)

    def _sheet_category_summary(self, ws):
        self._write_title(ws, 1, "Transaction Category Summary")
        headers = ["Category", "Credit Count", "Total Credits (₹)", "Debit Count", "Total Debits (₹)", "Net (₹)"]
        self._write_table_header(ws, 3, headers)

        from collections import defaultdict
        cat_data = defaultdict(lambda: {"cr_count": 0, "cr_total": 0.0, "dr_count": 0, "dr_total": 0.0})

        for txn in self.result.transactions:
            cat = txn.category or "Uncategorized"
            if txn.transaction_type == "credit":
                cat_data[cat]["cr_count"] += 1
                cat_data[cat]["cr_total"] += self._safe_float(txn.credit)
            elif txn.transaction_type == "debit":
                cat_data[cat]["dr_count"] += 1
                cat_data[cat]["dr_total"] += self._safe_float(txn.debit)

        for ri, (cat, data) in enumerate(sorted(cat_data.items()), 4):
            ws.row_dimensions[ri].height = 18
            net = data["cr_total"] - data["dr_total"]
            row_data = [cat, data["cr_count"], data["cr_total"], data["dr_count"], data["dr_total"], net]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = self.f_reg
                c.border = self.border_cell
                if ci in {3, 5, 6} and isinstance(val, float):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
            # Net column coloring
            net_c = ws.cell(row=ri, column=6)
            net_c.fill = self.fill_pass if net >= 0 else self.fill_alert
        self._autofit(ws)

    # ════════════════════════════════════════════════════════════
    # SHARED HELPERS
    # ════════════════════════════════════════════════════════════

    def _write_title(self, ws, row: int, title: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=title)
        c.font = self.f_title
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 32

    def _write_table_header(self, ws, row: int, columns: List[str]):
        ws.row_dimensions[row].height = 24
        for ci, col_name in enumerate(columns, 1):
            c = ws.cell(row=row, column=ci, value=col_name)
            c.font = self.f_header
            c.fill = self.fill_header
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = self.border_cell

    def _autofit(self, ws, min_width: int = 10, max_width: int = 55):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                if val.startswith("="):
                    val = "123,456.78"
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, max_width), min_width)

    def _to_float(self, val) -> Optional[float]:
        if val is None or val == "":
            return None
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, AttributeError):
            return val

    def _safe_float(self, val) -> float:
        try:
            return float(str(val or "0").replace(",", "").strip())
        except Exception:
            return 0.0
