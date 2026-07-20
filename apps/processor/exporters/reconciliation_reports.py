"""
Reconciliation Reports Exporter for CA Copilot.
Generates styled Excel, CSV, and PDF reports for Bank, GST, Ledger, Duplicates, and Exceptions.
"""
import csv
import io
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
from loguru import logger

class ReconciliationReportsExporter:
    def __init__(self, client_id: str, client_name: str):
        self.client_id = client_id
        self.client_name = client_name
        self._init_styles()

    def _init_styles(self):
        ff = "Segoe UI"
        self.f_title = Font(name=ff, size=16, bold=True, color="0F294A")
        self.f_subtitle = Font(name=ff, size=10, italic=True, color="475569")
        self.f_header = Font(name=ff, size=10, bold=True, color="FFFFFF")
        self.f_section = Font(name=ff, size=11, bold=True, color="0F294A")
        self.f_bold = Font(name=ff, size=9, bold=True, color="1E293B")
        self.f_reg = Font(name=ff, size=9, color="334155")
        
        self.fill_header = PatternFill(start_color="0F294A", end_color="0F294A", fill_type="solid")
        self.fill_section = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        self.fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        self.fill_match = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Light green
        self.fill_pending = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light amber
        self.fill_mismatch = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light red

        thin = Side(border_style="thin", color="CBD5E1")
        double = Side(border_style="double", color="1E293B")
        self.border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.border_total = Border(top=thin, bottom=double)

    # ── Excel Export ─────────────────────────────────────────────────────────

    def export_excel(self, data: Dict[str, Any], report_type: str, output_path: str) -> str:
        """Export reconciliation report to a styled Excel sheet."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reconciliation Summary"
        
        # Grid lines
        ws.views.sheetView[0].showGridLines = True

        if report_type == "bank":
            self._build_bank_excel(ws, data)
        elif report_type == "gst":
            self._build_gst_excel(ws, data)
        elif report_type == "ledger":
            self._build_ledger_excel(ws, data)
        elif report_type == "duplicate":
            self._build_duplicate_excel(ws, data)
        else:
            self._build_generic_excel(ws, data)

        # Autofit columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
        return output_path

    def _build_bank_excel(self, ws, data: Dict[str, Any]):
        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Bank Reconciliation Report — {self.client_name}"
        ws["A1"].font = self.f_title
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Offline Local Sync"
        ws["A2"].font = self.f_subtitle
        ws["A2"].alignment = Alignment(horizontal="center")

        # Stats Card
        ws["A4"] = "Reconciliation Summary Stats"
        ws["A4"].font = self.f_section
        
        ws["A5"] = "Total Bank Transactions"
        ws["B5"] = data.get("total_count", 0)
        ws["A6"] = "Matched Transactions"
        ws["B6"] = data.get("matched_count", 0)
        ws["A7"] = "Pending Review"
        ws["B7"] = data.get("pending_count", 0)
        ws["A8"] = "Unmatched Transactions"
        ws["B8"] = data.get("unmatched_count", 0)
        
        for r in range(5, 9):
            ws[f"A{r}"].font = self.f_bold
            ws[f"B{r}"].font = self.f_reg

        # Table Header
        headers = ["Date", "Narration", "Ref / Chq No", "Debit (Dr)", "Credit (Cr)", "Status", "Match Score / Reason"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_idx, value=h)
            cell.font = self.f_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Table Rows
        row_num = 11
        for txn in data.get("transactions", []):
            ws.cell(row=row_num, column=1, value=txn.get("date")).font = self.f_reg
            ws.cell(row=row_num, column=2, value=txn.get("narration")).font = self.f_reg
            ws.cell(row=row_num, column=3, value=txn.get("reference_number") or txn.get("cheque_number") or "").font = self.f_reg
            ws.cell(row=row_num, column=4, value=txn.get("debit")).font = self.f_reg
            ws.cell(row=row_num, column=5, value=txn.get("credit")).font = self.f_reg
            
            status_cell = ws.cell(row=row_num, column=6, value=txn.get("status", "").upper())
            status_cell.font = self.f_bold
            
            if txn.get("status") == "matched":
                status_cell.fill = self.fill_match
            elif txn.get("status") == "pending_review":
                status_cell.fill = self.fill_pending
            else:
                status_cell.fill = self.fill_mismatch

            reason_str = ""
            if txn.get("match_reason"):
                reason_str += f"[{txn.get('match_score')}%] {txn.get('match_reason')}"
            ws.cell(row=row_num, column=7, value=reason_str).font = self.f_reg

            for col_idx in range(1, 8):
                ws.cell(row=row_num, column=col_idx).border = self.border_cell

            row_num += 1

    def _build_gst_excel(self, ws, data: Dict[str, Any]):
        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"GST reconciliation (GSTR-2B vs Books) — {self.client_name}"
        ws["A1"].font = self.f_title
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = self.f_subtitle
        ws["A2"].alignment = Alignment(horizontal="center")

        # Table Header
        headers = ["Source", "GSTIN", "Vendor Name", "Invoice No", "Date", "Taxable Value", "GST Amount", "Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = self.f_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="left", vertical="center")

        row_num = 6
        for inv in data.get("invoices", []):
            ws.cell(row=row_num, column=1, value=inv.get("source_type", "").upper()).font = self.f_reg
            ws.cell(row=row_num, column=2, value=inv.get("gstin")).font = self.f_reg
            ws.cell(row=row_num, column=3, value=inv.get("vendor_name")).font = self.f_reg
            ws.cell(row=row_num, column=4, value=inv.get("invoice_number")).font = self.f_reg
            ws.cell(row=row_num, column=5, value=inv.get("invoice_date")).font = self.f_reg
            ws.cell(row=row_num, column=6, value=inv.get("taxable_value")).font = self.f_reg
            gst_tot = (inv.get("cgst") or 0.0) + (inv.get("sgst") or 0.0) + (inv.get("igst") or 0.0)
            ws.cell(row=row_num, column=7, value=gst_tot).font = self.f_reg
            
            status_cell = ws.cell(row=row_num, column=8, value=inv.get("status", "").upper())
            status_cell.font = self.f_bold
            if inv.get("status") == "matched":
                status_cell.fill = self.fill_match
            elif inv.get("status") == "pending_review":
                status_cell.fill = self.fill_pending
            else:
                status_cell.fill = self.fill_mismatch

            for col_idx in range(1, 9):
                ws.cell(row=row_num, column=col_idx).border = self.border_cell
            row_num += 1

    def _build_ledger_excel(self, ws, data: Dict[str, Any]):
        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Ledger Reconciliation Report — {self.client_name}"
        ws["A1"].font = self.f_title
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Ledger Type", "Date", "Description", "Ref / Invoice", "Debit", "Credit", "Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = self.f_header
            cell.fill = self.fill_header

        row_num = 5
        for entry in data.get("entries", []):
            ws.cell(row=row_num, column=1, value=entry.get("ledger_type", "").upper()).font = self.f_reg
            ws.cell(row=row_num, column=2, value=entry.get("date")).font = self.f_reg
            ws.cell(row=row_num, column=3, value=entry.get("description")).font = self.f_reg
            ws.cell(row=row_num, column=4, value=entry.get("reference_number") or entry.get("invoice_number") or "").font = self.f_reg
            ws.cell(row=row_num, column=5, value=entry.get("debit")).font = self.f_reg
            ws.cell(row=row_num, column=6, value=entry.get("credit")).font = self.f_reg
            
            status_cell = ws.cell(row=row_num, column=7, value=entry.get("status", "").upper())
            status_cell.font = self.f_bold
            if entry.get("status") == "matched":
                status_cell.fill = self.fill_match
            elif entry.get("status") == "pending_review":
                status_cell.fill = self.fill_pending
            else:
                status_cell.fill = self.fill_mismatch

            for col_idx in range(1, 8):
                ws.cell(row=row_num, column=col_idx).border = self.border_cell
            row_num += 1

    def _build_duplicate_excel(self, ws, data: Dict[str, Any]):
        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Duplicate Transactions & Invoice Report — {self.client_name}"
        ws["A1"].font = self.f_title
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Module", "Details", "Confidence", "Reason", "Record 1 ID", "Record 2 ID"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = self.f_header
            cell.fill = self.fill_header

        row_num = 5
        for m, dups in data.items():
            for d in dups:
                ws.cell(row=row_num, column=1, value=m.upper()).font = self.f_bold
                ws.cell(row=row_num, column=2, value=d.get("details")).font = self.f_reg
                
                conf_cell = ws.cell(row=row_num, column=3, value=d.get("confidence"))
                conf_cell.font = self.f_bold
                if d.get("confidence") == "Exact Match":
                    conf_cell.fill = self.fill_mismatch  # Red alert
                else:
                    conf_cell.fill = self.fill_pending

                ws.cell(row=row_num, column=4, value=d.get("reason")).font = self.f_reg
                ws.cell(row=row_num, column=5, value=d.get("record_1_id")).font = self.f_small
                ws.cell(row=row_num, column=6, value=d.get("record_2_id")).font = self.f_small

                for col_idx in range(1, 7):
                    ws.cell(row=row_num, column=col_idx).border = self.border_cell
                row_num += 1

    def _build_generic_excel(self, ws, data: Dict[str, Any]):
        ws["A1"] = "Reconciliation Report"
        ws["A1"].font = self.f_title
        ws["A3"] = str(data)

    # ── CSV Export ───────────────────────────────────────────────────────────

    def export_csv(self, data: Dict[str, Any], report_type: str) -> str:
        """Export reconciliation report to a CSV string."""
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type == "bank":
            writer.writerow(["Bank Reconciliation Report", f"Client: {self.client_name}"])
            writer.writerow([])
            writer.writerow(["Date", "Narration", "Ref / Cheque No", "Debit", "Credit", "Status", "Match Score", "Match Reason"])
            for txn in data.get("transactions", []):
                writer.writerow([
                    txn.get("date"),
                    txn.get("narration"),
                    txn.get("reference_number") or txn.get("cheque_number") or "",
                    txn.get("debit"),
                    txn.get("credit"),
                    txn.get("status"),
                    txn.get("match_score"),
                    txn.get("match_reason")
                ])
        elif report_type == "gst":
            writer.writerow(["GST Reconciliation Report (GSTR-2B vs Books)", f"Client: {self.client_name}"])
            writer.writerow([])
            writer.writerow(["Source", "GSTIN", "Vendor Name", "Invoice No", "Date", "Taxable Value", "CGST", "SGST", "IGST", "Status", "Match Reason"])
            for inv in data.get("invoices", []):
                writer.writerow([
                    inv.get("source_type"),
                    inv.get("gstin"),
                    inv.get("vendor_name"),
                    inv.get("invoice_number"),
                    inv.get("invoice_date"),
                    inv.get("taxable_value"),
                    inv.get("cgst"),
                    inv.get("sgst"),
                    inv.get("igst"),
                    inv.get("status"),
                    inv.get("match_reason")
                ])
        elif report_type == "ledger":
            writer.writerow(["Ledger Reconciliation Report", f"Client: {self.client_name}"])
            writer.writerow([])
            writer.writerow(["Ledger Type", "Date", "Description", "Ref / Invoice", "Debit", "Credit", "Status", "Match Reason"])
            for entry in data.get("entries", []):
                writer.writerow([
                    entry.get("ledger_type"),
                    entry.get("date"),
                    entry.get("description"),
                    entry.get("reference_number") or entry.get("invoice_number") or "",
                    entry.get("debit"),
                    entry.get("credit"),
                    entry.get("status"),
                    entry.get("match_reason")
                ])
        elif report_type == "duplicate":
            writer.writerow(["Duplicate Transactions Report", f"Client: {self.client_name}"])
            writer.writerow([])
            writer.writerow(["Module", "Details", "Confidence", "Reason", "Record 1 ID", "Record 2 ID"])
            for m, dups in data.items():
                for d in dups:
                    writer.writerow([
                        m.upper(),
                        d.get("details"),
                        d.get("confidence"),
                        d.get("reason"),
                        d.get("record_1_id"),
                        d.get("record_2_id")
                    ])
        else:
            writer.writerow(["Reconciliation Report", str(data)])

        return output.getvalue()

    # ── PDF Export ───────────────────────────────────────────────────────────

    def export_pdf(self, data: Dict[str, Any], report_type: str, output_path: str) -> str:
        """Export reconciliation report to a professional PDF using PyMuPDF."""
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        doc = fitz.open()

        page = doc.new_page()
        width, height = page.rect.width, page.rect.height

        # Title block
        page.draw_rect(fitz.Rect(0, 0, width, 80), fill=(15/255, 41/255, 74/255), color=None)
        page.insert_text(fitz.Point(30, 45), f"CA Copilot Reconciliation Report", fontsize=18, color=(1, 1, 1))
        page.insert_text(fitz.Point(30, 65), f"Client: {self.client_name} | Module: {report_type.upper()} | Generated: {datetime.now().strftime('%Y-%m-%d')}", fontsize=9, color=(0.8, 0.8, 0.8))

        y = 120
        page.insert_text(fitz.Point(30, y), f"Reconciliation Summaries & Findings", fontsize=12, color=(15/255, 41/255, 74/255))
        y += 20

        # Stats summary block
        if report_type == "bank":
            page.insert_text(fitz.Point(30, y), f"Total Transactions: {data.get('total_count', 0)}", fontsize=10)
            page.insert_text(fitz.Point(200, y), f"Matched: {data.get('matched_count', 0)}", fontsize=10)
            page.insert_text(fitz.Point(320, y), f"Pending: {data.get('pending_count', 0)}", fontsize=10)
            page.insert_text(fitz.Point(440, y), f"Unmatched: {data.get('unmatched_count', 0)}", fontsize=10)
            y += 40

            # List top 15 transactions
            page.insert_text(fitz.Point(30, y), "Detailed Transactions List (First 15):", fontsize=11, color=(15/255, 41/255, 74/255))
            y += 20

            headers = ["Date", "Narration", "Amount", "Status"]
            page.insert_text(fitz.Point(30, y), headers[0], fontsize=9)
            page.insert_text(fitz.Point(100, y), headers[1], fontsize=9)
            page.insert_text(fitz.Point(350, y), headers[2], fontsize=9)
            page.insert_text(fitz.Point(450, y), headers[3], fontsize=9)
            y += 15
            page.draw_line(fitz.Point(30, y), fitz.Point(550, y), color=(0.5, 0.5, 0.5))
            y += 15

            for txn in data.get("transactions", [])[:15]:
                if y > height - 50:
                    page = doc.new_page()
                    y = 50
                page.insert_text(fitz.Point(30, y), str(txn.get("date")), fontsize=8)
                page.insert_text(fitz.Point(100, y), str(txn.get("narration"))[:45], fontsize=8)
                amt = txn.get("debit") if txn.get("debit") > 0 else txn.get("credit")
                page.insert_text(fitz.Point(350, y), f"INR {amt:,.2f}", fontsize=8)
                status = str(txn.get("status")).upper()
                page.insert_text(fitz.Point(450, y), status, fontsize=8, color=(0, 0.5, 0) if status == "MATCHED" else (0.8, 0, 0))
                y += 18

        elif report_type == "gst":
            headers = ["GSTIN", "Invoice No", "Taxable Value", "Status"]
            page.insert_text(fitz.Point(30, y), headers[0], fontsize=9)
            page.insert_text(fitz.Point(180, y), headers[1], fontsize=9)
            page.insert_text(fitz.Point(300, y), headers[2], fontsize=9)
            page.insert_text(fitz.Point(420, y), headers[3], fontsize=9)
            y += 15
            page.draw_line(fitz.Point(30, y), fitz.Point(550, y), color=(0.5, 0.5, 0.5))
            y += 15

            for inv in data.get("invoices", [])[:20]:
                if y > height - 50:
                    page = doc.new_page()
                    y = 50
                page.insert_text(fitz.Point(30, y), str(inv.get("gstin")), fontsize=8)
                page.insert_text(fitz.Point(180, y), str(inv.get("invoice_number")), fontsize=8)
                page.insert_text(fitz.Point(300, y), f"INR {inv.get('taxable_value'):,.2f}", fontsize=8)
                status = str(inv.get("status")).upper()
                page.insert_text(fitz.Point(420, y), status, fontsize=8)
                y += 18
        else:
            page.insert_text(fitz.Point(30, y), f"Summary Details for {report_type.upper()}", fontsize=10)
            y += 20
            page.insert_text(fitz.Point(30, y), f"Check corresponding Excel or CSV formats for complete ledger breakdowns.", fontsize=9, color=(0.4, 0.4, 0.4))

        doc.save(output_path)
        doc.close()
        logger.info(f"PDF report saved: {output_path}")
        return output_path
